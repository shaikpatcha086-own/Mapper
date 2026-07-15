from openpyxl import Workbook


class ReportGenerator:

    def __init__(self):

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Mapping Report"

        headers = [
            "Target Field",
            "Source Field",
            "Confidence",
            "Method",
            "Status",
            "Reason"
        ]

        for col, header in enumerate(headers, start=1):
            self.sheet.cell(row=1, column=col).value = header

        self.current_row = 2

    def add_result(self, result):

        self.sheet.cell(self.current_row, 1).value = result["target_field"]
        self.sheet.cell(self.current_row, 2).value = result["source_field"]
        self.sheet.cell(self.current_row, 3).value = result["confidence"]
        self.sheet.cell(self.current_row, 4).value = result["method"]
        self.sheet.cell(self.current_row, 5).value = result["status"]
        self.sheet.cell(self.current_row, 6).value = result["reason"]

        self.current_row += 1

    def save(self, stream):

        self.workbook.save(stream)