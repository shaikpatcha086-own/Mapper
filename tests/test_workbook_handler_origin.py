import unittest
from io import BytesIO

from openpyxl import Workbook

from workbook_handler import WorkbookHandler


class TestWorkbookHandlerMappingOrigin(unittest.TestCase):

    def test_adds_mapping_origin_column_and_updates_value(self):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Field")
        ws.cell(row=1, column=2, value="source_field")
        ws.cell(row=2, column=1, value="CustomerAccount")

        payload = BytesIO()
        wb.save(payload)
        payload.seek(0)

        handler = WorkbookHandler(payload)

        handler.update_source_field(2, "ClientId")
        handler.update_mapping_origin(2, "CustTable")

        self.assertEqual(handler.sheet.cell(row=2, column=2).value, "ClientId")

        origin_header = handler.sheet.cell(row=1, column=3).value
        origin_value = handler.sheet.cell(row=2, column=3).value

        self.assertEqual(origin_header, "Mapping Source")
        self.assertEqual(origin_value, "CustTable")

    def test_uses_second_field_column_as_source_fallback(self):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Field")
        ws.cell(row=1, column=2, value="Field")
        ws.cell(row=2, column=1, value="CustomerAccount")
        ws.cell(row=2, column=2, value="")

        payload = BytesIO()
        wb.save(payload)
        payload.seek(0)

        handler = WorkbookHandler(payload)

        targets = handler.get_target_fields()
        self.assertEqual(len(targets), 1)
        self.assertEqual(targets[0]["field"], "CustomerAccount")

        handler.update_source_field(2, "ClientId")
        self.assertEqual(handler.sheet.cell(row=2, column=2).value, "ClientId")

    def test_processes_all_tabs_with_mapping_headers(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "READMEFIRST"
        ws1.cell(row=1, column=1, value="Notes")

        ws2 = wb.create_sheet("Template")
        ws2.cell(row=1, column=1, value="Field")
        ws2.cell(row=1, column=2, value="Field description")
        ws2.cell(row=1, column=3, value="source_field")
        ws2.cell(row=2, column=1, value="CustomerAccount")

        ws3 = wb.create_sheet("Customers V3")
        ws3.cell(row=1, column=1, value="Field")
        ws3.cell(row=1, column=2, value="Data Type")
        ws3.cell(row=1, column=3, value="Field")
        ws3.cell(row=2, column=1, value="InvoiceAccount")

        payload = BytesIO()
        wb.save(payload)
        payload.seek(0)

        handler = WorkbookHandler(payload)

        targets = handler.get_target_fields()
        self.assertEqual(len(targets), 2)
        self.assertEqual({x["sheet_name"] for x in targets}, {"Template", "Customers V3"})

        target_map = {x["field"]: x for x in targets}

        handler.update_source_field(
            target_map["CustomerAccount"]["row"],
            "ClientId",
            target_map["CustomerAccount"]["sheet_name"]
        )

        handler.update_source_field(
            target_map["InvoiceAccount"]["row"],
            "InvoiceCustomerAccount",
            target_map["InvoiceAccount"]["sheet_name"]
        )

        template_sheet = handler.workbook["Template"]
        customers_sheet = handler.workbook["Customers V3"]

        self.assertEqual(template_sheet.cell(row=2, column=3).value, "ClientId")
        self.assertEqual(customers_sheet.cell(row=2, column=3).value, "InvoiceCustomerAccount")


if __name__ == "__main__":
    unittest.main()
