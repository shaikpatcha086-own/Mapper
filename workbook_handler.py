"""
===========================================================
Workbook Handler
D365 Metadata Mapper V3
===========================================================

Purpose
-------
Reads the target metadata workbook while preserving all
formatting. Only updates the source_field column.
"""

from openpyxl import load_workbook
from copy import copy

from config import (
    TARGET_FIELD_HEADERS,
    TARGET_DESCRIPTION_HEADERS,
    TARGET_DATATYPE_HEADERS,
    TARGET_SOURCEFIELD_HEADERS,
    TARGET_MAPPING_ORIGIN_HEADERS,
    TARGET_MAPPING_ORIGIN_HEADER_NAME,
    TARGET_REQUIRED_HEADERS
)

from normalizer import normalize


class WorkbookHandler:

    def __init__(self, uploaded_file):

        uploaded_file.seek(0)

        self.workbook = load_workbook(uploaded_file)

        self.sheet = self.workbook.active

        self.mapping_sheets = []

        self.mapping_sheet_index = {}

        self.header_row = None

        self.field_col = None
        self.description_col = None
        self.source_field_col = None
        self.mapping_origin_col = None

        self._find_mapping_sheets()

    # ---------------------------------------------------------
    # Detect Header Row & Columns
    # ---------------------------------------------------------

    def _find_header_row(self, sheet):

        for row in sheet.iter_rows():
            for cell in row:
                value = normalize(cell.value)
                if value in TARGET_FIELD_HEADERS:
                    return cell.row

        return None

    def _ensure_mapping_origin_column(self, sheet_info):

        if sheet_info["mapping_origin_col"] is not None:
            return

        sheet = sheet_info["sheet"]

        mapping_origin_col = sheet.max_column + 1

        header_cell = sheet.cell(
            row=sheet_info["header_row"],
            column=mapping_origin_col
        )

        header_cell.value = TARGET_MAPPING_ORIGIN_HEADER_NAME

        source_header_cell = sheet.cell(
            row=sheet_info["header_row"],
            column=sheet_info["source_field_col"]
        )

        if source_header_cell.has_style:
            header_cell._style = copy(source_header_cell._style)

        if source_header_cell.number_format:
            header_cell.number_format = source_header_cell.number_format

        if source_header_cell.protection:
            header_cell.protection = copy(source_header_cell.protection)

        if source_header_cell.alignment:
            header_cell.alignment = copy(source_header_cell.alignment)

        sheet_info["mapping_origin_col"] = mapping_origin_col

    def _analyze_sheet(self, sheet):

        header_row = self._find_header_row(sheet)

        if header_row is None:
            return None

        field_columns = []
        field_col = None
        description_col = None
        source_field_col = None
        mapping_origin_col = None
        required_col = None
        has_datatype_header = False

        for cell in sheet[header_row]:

            value = normalize(cell.value)

            if value in TARGET_FIELD_HEADERS:
                field_columns.append(cell.column)

                if field_col is None:
                    field_col = cell.column

            elif value in TARGET_DESCRIPTION_HEADERS:
                description_col = cell.column

            elif value in TARGET_SOURCEFIELD_HEADERS:
                source_field_col = cell.column

            elif value in TARGET_MAPPING_ORIGIN_HEADERS:
                mapping_origin_col = cell.column

            elif value in TARGET_REQUIRED_HEADERS:
                required_col = cell.column

            elif value in TARGET_DATATYPE_HEADERS:
                has_datatype_header = True

        if field_col is None:
            return None

        if source_field_col is None and len(field_columns) >= 2:
            # Fallback for templates that repeat "Field" for source section.
            source_field_col = field_columns[-1]

        if source_field_col is None:
            return None

        signal_score = 1

        if description_col is not None:
            signal_score += 1

        if has_datatype_header:
            signal_score += 1

        info = {
            "sheet": sheet,
            "sheet_name": sheet.title,
            "header_row": header_row,
            "field_col": field_col,
            "description_col": description_col,
            "source_field_col": source_field_col,
            "mapping_origin_col": mapping_origin_col,
            "required_col": required_col,
            "signal_score": signal_score,
        }

        return info

    def _apply_primary_sheet_fields(self):

        primary = self.mapping_sheets[0]

        self.sheet = primary["sheet"]
        self.header_row = primary["header_row"]
        self.field_col = primary["field_col"]
        self.description_col = primary["description_col"]
        self.source_field_col = primary["source_field_col"]
        self.mapping_origin_col = primary["mapping_origin_col"]

    def _find_mapping_sheets(self):

        candidates = []

        for sheet in self.workbook.worksheets:
            info = self._analyze_sheet(sheet)
            if info is not None:
                candidates.append(info)

        if not candidates:
            raise Exception(
                "No mapping tab found. Expected headers: Field and Source_Field (or second Field column)."
            )

        # Prioritize stronger header signals but keep all valid mapping tabs.
        candidates.sort(
            key=lambda x: (
                x["signal_score"],
                x["sheet_name"].lower() == "template"
            ),
            reverse=True
        )

        for info in candidates:
            self._ensure_mapping_origin_column(info)

        self.mapping_sheets = candidates
        self.mapping_sheet_index = {
            x["sheet_name"]: x for x in self.mapping_sheets
        }

        self._apply_primary_sheet_fields()

    def _resolve_sheet_info(self, sheet_name):

        if sheet_name:
            info = self.mapping_sheet_index.get(sheet_name)
            if info is None:
                raise Exception(f"Unknown mapping sheet: {sheet_name}")
            return info

        return self.mapping_sheets[0]

    # ---------------------------------------------------------
    # Read Target Metadata
    # ---------------------------------------------------------

    def get_target_fields(self):

        metadata = []

        for sheet_info in self.mapping_sheets:

            sheet = sheet_info["sheet"]

            for row in range(sheet_info["header_row"] + 1,
                             sheet.max_row + 1):

                field = sheet.cell(
                    row=row,
                    column=sheet_info["field_col"]
                ).value

                if field is None:
                    continue

                field = str(field).strip()

                if field == "":
                    continue

                description = ""

                if sheet_info["description_col"]:

                    value = sheet.cell(
                        row=row,
                        column=sheet_info["description_col"]
                    ).value

                    if value:
                        description = str(value).strip()

                required = ""

                if sheet_info.get("required_col"):

                    req_value = sheet.cell(
                        row=row,
                        column=sheet_info["required_col"]
                    ).value

                    if req_value is not None:
                        required = str(req_value).strip().upper()

                metadata.append({

                    "row": row,

                    "field": field,

                    "description": description,

                    "required": required,

                    "sheet_name": sheet_info["sheet_name"]

                })

        return metadata

    # ---------------------------------------------------------
    # Update Source Field
    # ---------------------------------------------------------

    def update_source_field(
        self,
        row_number,
        source_field,
        sheet_name=None
    ):

        sheet_info = self._resolve_sheet_info(sheet_name)

        sheet_info["sheet"].cell(
            row=row_number,
            column=sheet_info["source_field_col"]
        ).value = source_field

    def update_mapping_origin(
        self,
        row_number,
        mapping_origin,
        sheet_name=None
    ):

        sheet_info = self._resolve_sheet_info(sheet_name)

        sheet_info["sheet"].cell(
            row=row_number,
            column=sheet_info["mapping_origin_col"]
        ).value = mapping_origin

    # ---------------------------------------------------------
    # Save Workbook
    # ---------------------------------------------------------

    def save(self, output_stream):

        self.workbook.save(output_stream)